from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render, redirect
from .models import Employee, Payslip
from .forms import EmployeeForm, ExcelUploadForm
from .filters import MonthFilter, EmployeeFilter
import pandas as pd
import math, csv, datetime
from django.db.models import F
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment


def input_employee_rates(request):
    if request.method == "POST":
        print("Post is done")
        form = EmployeeForm(request.POST)
        if form.is_valid():
            print("form is valid")
            form.save()
            return redirect("addemployee")

    else:
        print("Error")
        form = EmployeeForm()

    return render(request, "services/add_employee.html", {"form": form})


def emlist(request):
    sort_column = request.GET.get("sort", "emp_code")
    employee_filter = EmployeeFilter(request.GET, queryset=Employee.objects.all())

    if sort_column in ["emp_code", "basic"]:
        data = Employee.objects.all()
        data = sorted(data, key=lambda x: int(getattr(x, sort_column)))
    else:
        data = employee_filter.qs.order_by(sort_column)

    context = {
        "data": data,
        "employee_filter": employee_filter,
    }

    return render(request, "services/emlist.html", context)


def delete_employee(request, emp_code):
    if request.method == "POST":
        employee = get_object_or_404(Employee, emp_code=emp_code)
        employee.delete()
        return redirect(
            "emlist"
        )  # Replace 'employee_list' with the URL name of your employee list page
    else:
        return HttpResponseBadRequest("Invalid Request Method")


def upload_file(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            df = pd.read_excel(excel_file)

            employee_status = {}
            employee_total = {}
            current_emp_code = None
            for index, row in df.iterrows():
                emp_code = row["Emp_Code"]
                status = row["Status"]
                total = row["Total"]
                if not pd.isna(emp_code):
                    current_emp_code = emp_code
                    employee_status[current_emp_code] = []
                    employee_total[current_emp_code] = []
                if not pd.isna(status):
                    employee_status[current_emp_code].append(status)
                    employee_total[current_emp_code].append(total)

            present = {}
            absent = {}
            total_days = {}

            for i, j in employee_total.items():
                days_worked = 0
                absent_days = 0
                ot = 0
                wof_hours = 0
                total_days[i] = len(j)
                for k in j:
                        hours_worked = k.hour + k.minute / 60       
                        diff = hours_worked - 9
                        ot += max(math.floor(diff), 0)
                        # Check if the status on this day is "WO"
                        status_index = j.index(k)
                        if employee_status[i][status_index] == "WO":
                            wof_hours += max(
                                math.floor(hours_worked), 0
                            )  # Accumulate WOF hours
                        elif employee_status[i][status_index] == "A":
                            absent_days+=1
                        else:
                            if hours_worked > 9:
                                days_worked += 1

                present[i] = days_worked
                absent[i] = absent_days
                print(absent_days)
                
                employee = Employee.objects.get(emp_code=i)
                basicpay_perday = employee.basic_pay / 30
                basicpay_perhour = basicpay_perday / 24
                ot_amount = basicpay_perhour * ot

                wof_rate = basicpay_perhour * wof_hours

                total_earnings = (
                    employee.basic_pay + employee.sa + employee.hra + employee.pra_gain
                )

                gross_salary = (
                    total_earnings + employee.att_bonus + ot_amount + wof_rate
                )

                total_deductions = (
                    employee.pra_loss + employee.esi + employee.lop + employee.id_card
                )

                net_salary = gross_salary - total_deductions

                selected_month = form.cleaned_data.get("selected_month")

                selected_year = form.cleaned_data.get("year")

                payslip = Payslip.objects.create(
                    employee=employee,
                    total_days=total_days[i],
                    total_days_worked=present[i],
                    absent_days=(total_days[i] - present[i]),
                    WOF_hrs=wof_hours,
                    WOF_rate=wof_rate,
                    overtime_hrs=ot,
                    overtime_rate=ot_amount,
                    total_earnings=total_earnings,
                    gross_salary=gross_salary,
                    total_deductions=total_deductions,
                    net_salary=net_salary,
                    month=selected_month,
                    year=selected_year,
                )
                payslip.save()

            return redirect("payslip")
    else:
        form = ExcelUploadForm()

    return render(request, "services/upload.html", {"form": form})


def payslip(request):
    payslip = MonthFilter(data=request.GET, queryset=Payslip.objects.all())

    # Extract the selected month and year from the filter form
    selected_month = request.GET.get("month")
    selected_year = request.GET.get("year")

    return render(
        request,
        "services/payslip.html",
        {
            "filter": payslip,
            "selected_month": selected_month,
            "selected_year": selected_year,
        },
    )


def profile(request, user_id):
    employee = get_object_or_404(Employee, emp_code=user_id)
    payslips = Payslip.objects.filter(employee=employee)
    context = {"employee": employee, "payslips": payslips}
    return render(request, "services/profile.html", context)


def export_salary_summary(request, selected_month, selected_year):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=payslip.csv"

    writer = csv.writer(response)

    payslips = Payslip.objects.filter(month=selected_month, year=selected_year)

    writer.writerow(
        [
            "Name",
            "ID",
            "Month",
            "Days worked",
            "OT hours",
            "Basic",
            "SA",
            "HRA",
            "PRA_gain",
            "Total",
            "OT Amt.",
            "ATTD. Bonus",
            "Gross Salary",
            "LOP",
            "ESI",
            "PRA_gain",
            "ID CARD",
            "Total",
            "Net Salary",
        ]
    )

    for payslip in payslips:
        writer.writerow(
            [
                payslip.employee.emp_name,
                payslip.employee.emp_code,
                payslip.month,
                payslip.total_days_worked,
                payslip.overtime_hrs,
                payslip.employee.basic_pay,
                payslip.employee.sa,
                payslip.employee.hra,
                payslip.employee.pra_gain,
                payslip.total_earnings,
                payslip.overtime_rate,
                payslip.employee.att_bonus,
                payslip.gross_salary,
                payslip.employee.lop,
                payslip.employee.esi,
                payslip.employee.pra_loss,
                payslip.employee.id_card,
                payslip.total_deductions,
                payslip.net_salary,
            ]
        )

    return response


def export_payslip(request, selected_month, selected_year):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    blue_fill = PatternFill(start_color="6B6BEF", end_color="6B6BEF", fill_type="solid")
    green_fill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )

    # Query all payslips of Payslip
    payslips = Payslip.objects.filter(month=selected_month, year=selected_year)

    centered_alignment = Alignment(horizontal="center", vertical="center")

    merge_rows_AB_DE = ["Earning", "Authorized Signatory", " "]
    merge_rows_BC = ["Employee Name", "Employee Code", "Designation", "Department"]

    # Iterate over each Payslip payslip and add data to the list
    for payslip in payslips:
        common_data = [
            ["Ministon Engineering Services"],
            ["EAST POTHERI, KATTANKULATHUR, CHENNAI - 603 203"],
            [f"Salary Slip for the month : {payslip.month} {payslip.year}"],
        ]

        for i, row in enumerate(common_data):
            cell = ws.cell(row=ws.max_row + 1, column=1, value=row[0])
            cell.font = cell.font = Font(
                name="Lucida Console", size=20, bold=True, italic=False, color="000000"
            )
            cell.alignment = centered_alignment
            if i == 2:  # Check if it's the third row (index 2) and apply the green fill
                cell.fill = green_fill
            else:
                cell.fill = blue_fill
            ws.merge_cells(
                start_row=cell.row, start_column=1, end_row=cell.row, end_column=5
            )

        data = [
            [
                "Employee Name",
                payslip.employee.emp_name,
                "",
                "Paid Days",
                payslip.total_days_worked,
            ],
            [
                "Employee Code",
                payslip.employee.emp_code,
                "",
                "LOPs",
                payslip.absent_days,
            ],
            [
                "Designation",
                payslip.employee.department,
                "",
                "OT hours",
                payslip.overtime_hrs,
            ],
            [
                "Department",
                payslip.employee.department,
                "",
                "WOF hours",
                payslip.WOF_hrs,
            ],
            ["Total Days", payslip.total_days],
            [" "],
            ["Earning", "", "", "Deductions", ""],
            [
                "Basic",
                f"₹{payslip.employee.basic_pay}",
                "",
                "LOP",
                f"₹{payslip.employee.lop}",
            ],
            [
                "SA",
                f"₹{payslip.employee.sa}",
                "",
                "ESI/Health Insurance",
                f"₹{payslip.employee.esi}",
            ],
            [
                "HRA",
                f"₹{payslip.employee.hra}",
                "",
                "PRA",
                f"₹{payslip.employee.pra_loss}",
            ],
            [
                "PRA",
                f"₹{payslip.employee.pra_gain}",
                "",
                "Loan Recovery",
                f"₹{payslip.employee.id_card}",
            ],
            ["WOF", f"₹{payslip.WOF_rate}"],
            ["OT", f"₹{payslip.overtime_rate}"],
            ["Attd. Bonus", f"₹{payslip.employee.att_bonus}"],
            [
                "Gross Earnings",
                f"₹{payslip.gross_salary}",
                "",
                "Total Deductions",
                f"₹{payslip.total_deductions}",
            ],
            [" "],
            ["Net Payable", "", f"₹{payslip.net_salary}", "", ""],
            [" "],
            [" "],
            [" "],
            ["Authorized Signatory", "", "", "Employee Signatory", ""],
            [" "],
        ]

        # Iterate over data rows and merge columns B and C if the row content matches the criteria
        for i, row in enumerate(data):
            ws.append(row)
            if row[0] in merge_rows_AB_DE:
                ws.merge_cells(
                    start_row=ws.max_row,
                    start_column=1,
                    end_row=ws.max_row,
                    end_column=2,
                )
                ws.merge_cells(
                    start_row=ws.max_row,
                    start_column=4,
                    end_row=ws.max_row,
                    end_column=5,
                )
            elif row[0] in merge_rows_BC:
                ws.merge_cells(
                    start_row=ws.max_row,
                    start_column=2,
                    end_row=ws.max_row,
                    end_column=3,
                )
            elif row[0] in ["Net Payable"]:
                ws.merge_cells(
                    start_row=ws.max_row,
                    start_column=1,
                    end_row=ws.max_row,
                    end_column=2,
                )
                ws.merge_cells(
                    start_row=ws.max_row,
                    start_column=3,
                    end_row=ws.max_row,
                    end_column=4,
                )

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Apply the custom font style to the entire workbook
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in [
                    "Earning",
                    "Deductions",
                    "Net Payable",
                ]:
                    cell.font = Font(
                        name="Lucida Console",
                        size=12,
                        bold=True,
                        italic=False,
                        color="000000",
                    )
                else:
                    cell.font = Font(
                        name="Lucida Console",
                        size=12,
                        bold=False,
                        italic=False,
                        color="000000",
                    )

                if cell.value in [
                    "Authorized Signatory",
                    "Employee Signatory",
                    "Net Payable",
                ]:
                    cell.alignment = centered_alignment
                elif cell.value in [
                    payslip.net_salary,
                    payslip.employee.emp_code,
                    payslip.employee.basic_pay,
                    payslip.employee.sa,
                    payslip.employee.hra,
                    payslip.employee.pra_gain,
                    payslip.employee.pra_loss,
                    payslip.overtime_rate,
                    payslip.overtime_hrs,
                    payslip.total_days_worked,
                    payslip.employee.att_bonus,
                    payslip.gross_salary,
                    payslip.total_deductions,
                    payslip.absent_days,
                    payslip.total_days,
                    payslip.employee.lop,
                    payslip.employee.esi,
                    payslip.employee.id_card,
                    payslip.WOF_hrs,
                    payslip.WOF_rate
                ]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Reduce the width of columns for data rows (adjust as needed)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 20

    # Customize the Excel file design (e.g., font, colors, etc.)
    # Example: Changing font style and size for the header row

    # Save the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=Payslip.xlsx"
    wb.save(response)

    return response
